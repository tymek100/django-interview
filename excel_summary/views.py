import io
import ast
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from rest_framework import parsers, status
from rest_framework.response import Response
from rest_framework.views import APIView

from drf_spectacular.utils import (
    extend_schema,
    OpenApiResponse,
)

from .serializers import (
    ExcelSummaryRequestSerializer,
    ExcelSummaryResponseSerializer,
)


def detect_header_row(ws, requested_columns: List[str], max_header_search: int = 5):
    """
    Try to detect which row contains the headers by matching requested column
    names against the first `max_header_search` rows.

    Returns: (row_index, row_values) or (None, None) if nothing found.
    """
    requested_norm = [normalize_header(c) for c in requested_columns]

    best_row_idx: Optional[int] = None
    best_match_count = 0
    best_row_values = None

    # 1) Try to find the row where most requested column names appear
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_header_search, values_only=True),
        start=1,
    ):
        if not any(cell not in (None, "") for cell in row):
            # skip completely empty rows
            continue

        header_map = {
            normalize_header(cell): col_idx
            for col_idx, cell in enumerate(row)
            if normalize_header(cell)
        }

        match_count = sum(1 for c in requested_norm if c in header_map)

        if match_count > best_match_count:
            best_match_count = match_count
            best_row_idx = idx
            best_row_values = row

    if best_row_idx is not None and best_match_count > 0:
        # Found a row that matches at least one requested column
        return best_row_idx, best_row_values

    # 2) Fallback: first non-empty row
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_header_search, values_only=True),
        start=1,
    ):
        if any(cell not in (None, "") for cell in row):
            return idx, row

    return None, None


def normalize_header(header: Any) -> str:
    """Normalize header cell to allow case-insensitive matching."""
    if header is None:
        return ""
    return str(header).strip().lower()


def coerce_to_number(value: Any) -> Optional[float]:
    """
    Try to convert various cell values to float.

    Handles:
    - numeric types
    - strings like "$90,00", "90,00", "90.00", "1,234.56"
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None

        # Remove common currency symbols and spaces
        for ch in ["$", "€", "£"]:
            s = s.replace(ch, "")
        s = s.replace(" ", "")

        # Handle decimal/thousands separators
        if "," in s and "." in s:
            # Assume '.' is decimal, ',' is thousands separator: "1,234.56"
            s = s.replace(",", "")
        elif "," in s and "." not in s:
            # European style: "90,00" -> "90.00"
            s = s.replace(",", ".")

        try:
            return float(s)
        except ValueError:
            return None

    return None


class ExcelSummaryView(APIView):
    """
    POST /api/excel-summary/

    multipart/form-data:
        file: <Excel file>
        columns: <column name>   (can appear multiple times)

    Response:
    {
        "file": "<filename>",
        "summary": [
            {"column": "CURRENT USD", "sum": 1234.5, "avg": 56.7},
            ...
        ]
    }
    """

    parser_classes = [parsers.MultiPartParser, parsers.FormParser]

    @extend_schema(
        tags=["Excel Upload"],
        summary="Summarize numeric columns in an Excel file",
        description=(
            "Upload an Excel `.xlsx` file and provide a list of column header names.\n\n"
            "The API will locate each column and return the sum and average of "
            "all numeric values found in that column."
        ),
        request=ExcelSummaryRequestSerializer,
        responses={
            200: ExcelSummaryResponseSerializer,
            400: OpenApiResponse(
                description="Validation error or Excel parsing error."
            ),
        },
    )
    def post(self, request, *args, **kwargs):
        request_serializer = ExcelSummaryRequestSerializer(data=request.data)
        request_serializer.is_valid(raise_exception=True)

        uploaded_file = request_serializer.validated_data["file"]
        columns = ast.literal_eval(request_serializer.validated_data["columns"])

        # Read workbook from in-memory bytes
        try:
            file_bytes = io.BytesIO(uploaded_file.read())
            wb = load_workbook(filename=file_bytes, data_only=True)
        except Exception:
            return Response(
                {"detail": "Unable to read Excel file. Make sure it is a valid .xlsx file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active  # use first sheet

        # Find header row based on requested column names
        header_row_index, header_row_values = detect_header_row(ws, columns)

        if header_row_index is None or header_row_values is None:
            return Response(
                {"detail": "Could not detect header row in the Excel sheet."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if header_row_index is None or header_row_values is None:
            return Response(
                {"detail": "Could not detect header row in the Excel sheet."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Map normalized header -> column index
        header_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(header_row_values):
            key = normalize_header(cell)
            if key:
                header_map[key] = col_idx

        requested_summaries = []
        missing_columns = []

        # Iterate over requested columns and compute sum + avg
        for col_name in columns:
            norm = normalize_header(col_name)
            if norm not in header_map:
                missing_columns.append(col_name)
                continue

            col_idx = header_map[norm]
            total = 0.0
            count = 0

            # Iterate data rows, starting after header
            for row in ws.iter_rows(
                min_row=header_row_index + 1,
                max_row=ws.max_row,
                values_only=True,
            ):
                if col_idx >= len(row):
                    continue
                cell_value = row[col_idx]
                number = coerce_to_number(cell_value)
                if number is None:
                    continue
                total += number
                count += 1

            if count > 0:
                avg = total / count
                requested_summaries.append(
                    {
                        "column": col_name,
                        "sum": round(total, 2),
                        "avg": round(avg, 2),
                    }
                )
            else:
                # Column exists but no numeric values
                requested_summaries.append(
                    {
                        "column": col_name,
                        "sum": 0.0,
                        "avg": 0.0,
                    }
                )

        if not requested_summaries:
            return Response(
                {
                    "detail": "None of the requested columns were found in the sheet.",
                    "requested_columns": columns,
                    "available_columns": list(header_map.keys()),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        response_data = {
            "file": uploaded_file.name,
            "summary": requested_summaries,
        }

        response_serializer = ExcelSummaryResponseSerializer(data=response_data)
        response_serializer.is_valid(raise_exception=True)

        return Response(response_serializer.data, status=status.HTTP_200_OK)
